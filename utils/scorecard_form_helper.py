import os
import base64
import tempfile
from datetime import datetime
from dotenv import load_dotenv
from urllib.parse import quote
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from utils.airtable_throttler import AirtableThrottler

load_dotenv()
airtable_requester = AirtableThrottler()


def _create_kpi_checklist_excel(kpi_checklist_fieldset, employee_position=""):
    items = kpi_checklist_fieldset.get("items", {}) if isinstance(kpi_checklist_fieldset, dict) else {}
    if not items:
        return None

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KPI Checklist"
    sheet.sheet_view.showGridLines = False

    headers = ["KPI Description", "Expectations", "is the KPI met?", "Remarks/Proof"]
    remarks_col_index = headers.index("Remarks/Proof") + 1
    sheet.append(headers)

    header_fill = PatternFill(start_color="009644", end_color="009644", fill_type="solid")
    thin_side = Side(style="thin", color="A6A6A6")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_index in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_index)
        cell.fill = header_fill
        cell.border = cell_border
        cell.alignment = Alignment(vertical="center")
        cell.font = Font(bold=True, color="FFFFFF")

    def sort_key(item_key):
        return int(item_key) if str(item_key).isdigit() else str(item_key)

    def normalize_links(raw_links):
        if raw_links is None:
            return ""
        # Convert multiline links into a clean single-line value for Excel.
        lines = str(raw_links).replace("\r\n", "\n").replace("\r", "\n").split("\n")
        cleaned_lines = [line.strip() for line in lines if line.strip()]
        return "\n".join(cleaned_lines)

    for item_no in sorted(items.keys(), key=sort_key):
        item = items.get(item_no, {}) or {}
        sheet.append([
            item.get("description", ""),
            item.get("expectation", ""),
            "Yes" if item.get("met", False) else "No",
            normalize_links(item.get("links", ""))
        ])

        # Keep each link on its own line in the same cell.
        sheet.cell(row=sheet.max_row, column=remarks_col_index).alignment = Alignment(
            vertical="center",
            wrap_text=True
        )

    for row_index in range(2, sheet.max_row + 1):
        for col_index in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.border = cell_border
            if col_index != remarks_col_index:
                cell.alignment = Alignment(vertical="center")

    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 23.4

    # Auto-size columns with sane min/max width so long text does not break layout.
    for col_index in range(1, len(headers) + 1):
        max_length = 0
        for row_index in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            if cell_value is None:
                continue
            max_length = max(max_length, len(str(cell_value)))

        adjusted_width = min(max(max_length + 2, 12), 80)
        sheet.column_dimensions[get_column_letter(col_index)].width = adjusted_width

    safe_record_id = employee_position or "unknown_record"
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"kpi_checklist_{safe_record_id}_{timestamp}.xlsx"
    output_path = os.path.join(tempfile.gettempdir(), filename)
    workbook.save(output_path)
    return output_path


def _upload_kpi_excel_to_airtable(record_id, file_path):
    attachment_field_id = "fldXevnZ5KL8myL4G"
    airtable_api_key = os.getenv("AIRTABLE_API_KEY", "")
    if not airtable_api_key:
        print("Skipping KPI Excel attachment upload: AIRTABLE_API_KEY is not set")
        return False

    if not record_id or not file_path or not os.path.exists(file_path):
        print("Skipping KPI Excel attachment upload: missing record id or file")
        return False

    with open(file_path, "rb") as file_stream:
        encoded_file = base64.b64encode(file_stream.read()).decode("utf-8")

    upload_url = f"https://content.airtable.com/v0/appVBupdRP0pwHBjh/{record_id}/{attachment_field_id}/uploadAttachment"
    payload = {
        "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "filename": os.path.basename(file_path),
        "file": encoded_file
    }

    response = requests.post(
        upload_url,
        headers={
            "Authorization": f"Bearer {airtable_api_key}",
            "Content-Type": "application/json"
        },
        json=payload,
        timeout=60
    )

    if response.status_code not in [200, 201]:
        print(f"Failed KPI Excel attachment upload: {response.status_code} - {response.text}")
        return False

    print("KPI checklist Excel uploaded successfully")
    return True

def get_user(rec_id):
    print("Retrieving workers from Airtable")
    try:
        response = airtable_requester.throttled_get(
            url=f'https://api.airtable.com/v0/appfccXiah8EtMfbZ/tblU2uvcUVqERiRzv/{rec_id}',
            headers = {
                "Authorization": "Bearer " + os.getenv('AIRTABLE_API_KEY'),
                "Content-Type": "application/json",
            }
        )
        response_data = response.json()
        records = response_data.get('fields', [])
    except Exception as e:
        print(f"Error retrieving Airtable records: {e}")
        records = []
    return records

def get_user_by(field="Record ID", value=None):
    print(f"Retrieving user by {field}: {value}")
    try:
        url = f'https://api.airtable.com/v0/appVBupdRP0pwHBjh/tblMrivRZBk0ZAJbK'
        response = airtable_requester.throttled_get(
            url=url,
            headers = {
                "Authorization": "Bearer " + os.getenv('AIRTABLE_API_KEY'),
                "Content-Type": "application/json",
            },
            params={
                "filterByFormula": f'FIND("{value}", {{{field}}})'
            }
        )
        response_data = response.json()
        record_id = response_data.get('records', [{}])[0].get('id', None)
    except Exception as e:
        print(f"Error retrieving Airtable records: {e}")
        record_id = None
    return record_id

def get_kpi_checklist_fields(position):
    print(f"Retrieving KPI checklist fields for position: {position}")
    try:
        url = f'https://api.airtable.com/v0/appVBupdRP0pwHBjh/tblsEjOXdMsKdFNaW?filterByFormula=' + quote(f'FIND("{position}", {{Position Title}})')
        response = airtable_requester.throttled_get(
            url=url,
            headers = {
                "Authorization": "Bearer " + os.getenv('AIRTABLE_API_KEY'),
                "Content-Type": "application/json",
            }
        )
        response_data = response.json()
        records = response_data.get('records', [])
        
        form_schema = []
        for record in records:
            fields = record.get('fields', {})
            record_id = record.get('id')
            kpi_name = fields.get('KPI Description')
            expectation = fields.get('Expectations')
            is_required = fields.get('Is Required', False)
            if kpi_name:
                form_schema.append({
                    "name": record_id,
                    "description": kpi_name,
                    "expectation": expectation,
                    "isRequired": is_required
                })
        final_schema = {
        "section": "KPI Checklist",
        "fields": form_schema
        }

    except Exception as e:
        print(f"Error retrieving KPI checklist fields: {e}")
        final_schema = {
            "section": "KPI Checklist",
            "fields": []
        }
    return final_schema


def submit_data_to_airtable(data):
    print(data)
    print("Submitting data to Airtable")
    fields = {}

    # Standard fields
    for key1, value in data.items():
        #deny sections for custom fields
        if key1 in ["kpi_checklist_fieldset", "recordId"]: continue

        if "items" in value:
            for key2, item in value["items"].items():
                fields[item["label"]] = item.get("value", "")
        else:
            for key2, item in value.items():
                if key2 in ["fullname", "email", "record_id"]: continue
                fields[key2] = item

    
    #Custom fields
    employee_name = data.get("employee_being_scored_fieldset", {}).get("fullname", "")
    employee_position = data.get("employee_being_scored_fieldset", {}).get("position", "")
    proctor = data.get("scorecard_proctor_fieldset", {})
    proctor_id = proctor.get("record_id", "")
    proctor_email = proctor.get("email", "")
    proctor_data = get_user_by(field="Record ID", value=proctor_id) if proctor_id else get_user_by(field="Work Email Address Copy", value=proctor_email)
    fields["scorecard_proctor"] = [proctor_data]
    fields["employee"] = [get_user_by(field="Record ID", value=data.get("recordId", {}))]

    employee_name = data.get("employee_being_scored_fieldset", {}).get("fullname", "")
    employee_position = data.get("employee_being_scored_fieldset", {}).get("position", "")
    kpi_excel_path = _create_kpi_checklist_excel(
        data.get("kpi_checklist_fieldset", {}),
        f"{employee_name}_{employee_position}"
    )

    try:
        response = airtable_requester.throttled_post(
            url='https://api.airtable.com/v0/appVBupdRP0pwHBjh/tblQyRrzoGOVLve2a',
            headers = {
                "Authorization": "Bearer " + os.getenv('AIRTABLE_API_KEY'),
                "Content-Type": "application/json",
            },
            json={
                "fields": fields
            }
        )
        if response.status_code == 200:
            print("Data submitted successfully")

            created_record_id = response.json().get("id", "")
            if kpi_excel_path:
                is_upload_success = _upload_kpi_excel_to_airtable(created_record_id, kpi_excel_path)
                if not is_upload_success:
                    return False

            return True
        else:
            print(f"Failed to submit data: {response.status_code} - {response.text}")
            return False
    except Exception as e:
        print(f"Error submitting data to Airtable: {e}")
        return False
    finally:
        if kpi_excel_path and os.path.exists(kpi_excel_path):
            os.remove(kpi_excel_path)